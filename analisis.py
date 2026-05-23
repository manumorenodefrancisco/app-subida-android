import sqlite3
import math
import os
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def calcular_nivel(xp):
    return 0 if xp <= 0 else int(math.sqrt(xp / 100))


def xp_para_nivel(nivel):
    return nivel * nivel * 100


def progreso_al_siguiente(xp):
    nivel = calcular_nivel(xp)
    base = xp_para_nivel(nivel)
    siguiente = xp_para_nivel(nivel + 1) - base
    return 0 if siguiente <= 0 else round(((xp - base) / siguiente) * 100, 1)


def asegurar_columnas(df, defaults):
    if df is None:
        df = pd.DataFrame()
    for col, valor in defaults.items():
        if col not in df.columns:
            df[col] = valor
    return df


def leer_bd(ruta):
    con = sqlite3.connect(ruta)

    try:
        usuario = pd.read_sql("SELECT * FROM usuario LIMIT 1", con)
    except:
        usuario = pd.DataFrame()

    try:
        tareas = pd.read_sql("SELECT * FROM tareas", con)
    except:
        tareas = pd.DataFrame()

    try:
        registros = pd.read_sql("SELECT * FROM registro_sistema", con)
    except:
        registros = pd.DataFrame()

    try:
        progreso = pd.read_sql("SELECT * FROM progreso_diario", con)
    except:
        progreso = pd.DataFrame()

    con.close()

    usuario = asegurar_columnas(usuario, {
        "nombre": "Sin datos", "xpTotal": 0, "monedas": 0, "rachaDias": 0, "tituloActivo": "-"
    })
    tareas = asegurar_columnas(tareas, {
        "id": 0, "completada": 0, "dificultad": "Sin datos", "xpRecompensa": 0, "monedasRecompensa": 0
    })
    registros = asegurar_columnas(registros, {
        "tipo": "Sin datos", "mensaje": "", "fechaHora": "", "leido": 0
    })
    progreso = asegurar_columnas(progreso, {
        "fecha": "", "tareasCompletadas": 0, "xpGanado": 0, "monedasGanadas": 0, "pasos": 0
    })

    for col in ["xpTotal", "monedas", "rachaDias"]:
        usuario[col] = pd.to_numeric(usuario[col], errors="coerce").fillna(0)

    for col in ["completada", "xpRecompensa", "monedasRecompensa"]:
        tareas[col] = pd.to_numeric(tareas[col], errors="coerce").fillna(0)

    for col in ["leido"]:
        registros[col] = pd.to_numeric(registros[col], errors="coerce").fillna(0)

    for col in ["tareasCompletadas", "xpGanado", "monedasGanadas", "pasos"]:
        progreso[col] = pd.to_numeric(progreso[col], errors="coerce").fillna(0)

    if len(usuario) == 0:
        usuario = pd.DataFrame([{
            "nombre": "Sin datos", "xpTotal": 0, "monedas": 0, "rachaDias": 0, "tituloActivo": "-"
        }])

    return usuario, tareas, registros, progreso


def generar_excel(ruta_bd, ruta_salida):
    usuario, tareas, registros, progreso = leer_bd(ruta_bd)

    u = usuario.iloc[0]
    xp = int(u["xpTotal"])
    nivel = calcular_nivel(xp)

    tareas_completadas = tareas[tareas["completada"] == 1]
    tareas_pendientes = tareas[tareas["completada"] == 0]
    pct = round(len(tareas_completadas) / len(tareas) * 100, 1) if len(tareas) else 0

    writer = pd.ExcelWriter(ruta_salida, engine="openpyxl")

    pd.DataFrame({
        "Campo": [
            "Nombre", "Nivel actual", "XP total", "Progreso al siguiente nivel",
            "XP para el siguiente nivel", "Monedas", "Racha de días", "Título activo",
            "Total tareas creadas", "Tareas completadas", "Tareas pendientes", "% Completadas"
        ],
        "Valor": [
            u["nombre"], nivel, xp, f"{progreso_al_siguiente(xp)}%",
            xp_para_nivel(nivel + 1), int(u["monedas"]), int(u["rachaDias"]),
            u["tituloActivo"], len(tareas), len(tareas_completadas),
            len(tareas_pendientes), f"{pct}%"
        ]
    }).to_excel(writer, sheet_name="Jugador", index=False)

    if len(tareas_completadas):
        por_dificultad = tareas_completadas.groupby("dificultad").agg(
            completadas=("id", "count"),
            xp_ganada=("xpRecompensa", "sum"),
            monedas_ganadas=("monedasRecompensa", "sum")
        ).reset_index()
        por_dificultad.columns = ["Dificultad", "Completadas", "XP ganada", "Monedas ganadas"]
    else:
        por_dificultad = pd.DataFrame(columns=["Dificultad", "Completadas", "XP ganada", "Monedas ganadas"])

    por_dificultad.to_excel(writer, sheet_name="Tareas por dificultad", index=False)

    progreso_sorted = progreso.sort_values("fecha").copy()
    progreso_sorted["xp_acumulada"] = progreso_sorted["xpGanado"].cumsum() if len(progreso_sorted) else pd.Series(dtype=float)
    progreso_sorted["nivel_ese_dia"] = progreso_sorted["xp_acumulada"].apply(calcular_nivel) if len(progreso_sorted) else pd.Series(dtype=float)

    export_progreso = progreso_sorted[[
        "fecha", "tareasCompletadas", "xpGanado", "monedasGanadas", "pasos", "xp_acumulada", "nivel_ese_dia"
    ]].copy()
    export_progreso.columns = ["Fecha", "Tareas", "XP ganada", "Monedas", "Pasos", "XP acumulada", "Nivel ese día"]
    export_progreso.to_excel(writer, sheet_name="Progreso diario", index=False)

    reg_export = registros[["tipo", "mensaje", "fechaHora", "leido"]].copy()
    reg_export["leido"] = reg_export["leido"].map({0: "No", 1: "Sí"}).fillna("No")
    reg_export.columns = ["Tipo", "Mensaje", "Fecha y hora", "Leído"]
    if len(reg_export):
        reg_export = reg_export.sort_values("Fecha y hora", ascending=False)
    reg_export.to_excel(writer, sheet_name="Registro sistema", index=False)

    xp_media = progreso_sorted["xpGanado"].mean() if len(progreso_sorted) else 25
    if pd.isna(xp_media):
        xp_media = 25

    pd.DataFrame([
        {"Horizonte": f"En {d} días", "XP proyectada": int(xp + xp_media * d), "Nivel proyectado": calcular_nivel(int(xp + xp_media * d))}
        for d in [7, 14, 30, 60, 90]
    ]).assign(**{
        "Niveles que subirías": lambda df: df["Nivel proyectado"] - nivel
    }).to_excel(writer, sheet_name="Proyección", index=False)

    writer.close()

    wb = load_workbook(ruta_salida)
    aplicar_estilos(wb)
    wb.save(ruta_salida)


def aplicar_estilos(wb):
    cabecera_fill = PatternFill("solid", fgColor="0D0D2B")
    cabecera_font = Font(bold=True, color="00D4FF", size=11)
    fila_par_fill = PatternFill("solid", fgColor="E8F4FF")
    fila_impar_fill = PatternFill("solid", fgColor="FFFFFF")
    centro = Alignment(horizontal="center", vertical="center")

    for nombre in wb.sheetnames:
        ws = wb[nombre]

        for celda in ws[1]:
            if celda.value is not None:
                celda.fill = cabecera_fill
                celda.font = cabecera_font
                celda.alignment = centro

        for i, fila in enumerate(ws.iter_rows(min_row=2), start=2):
            relleno = fila_par_fill if i % 2 == 0 else fila_impar_fill
            for celda in fila:
                celda.fill = relleno
                celda.alignment = centro

        for col in ws.columns:
            ancho = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 45)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Informe de datos")
        self.root.geometry("480x300")
        self.root.resizable(False, False)
        self.root.configure(bg="#0D0D1A")

        self.ruta_bd = tk.StringVar(value="Sin seleccionar")
        self.ruta_salida = tk.StringVar(value="Sin seleccionar")

        tk.Label(root, text="Generador de informes", font=("Courier", 13, "bold"),
                 fg="#00D4FF", bg="#0D0D1A").pack(pady=(18, 2))
        tk.Label(root, text="Carga la base de datos exportada desde el dispositivo",
                 font=("Courier", 8), fg="#4A6080", bg="#0D0D1A").pack(pady=(0, 14))

        self.fila("Base de datos (.db):", self.ruta_bd, self.elegir_bd)
        self.fila("Guardar Excel en:", self.ruta_salida, self.elegir_salida)

        self.lbl_estado = tk.Label(root, text="", font=("Courier", 9),
                                   fg="#00FF88", bg="#0D0D1A")
        self.lbl_estado.pack(pady=8)

        tk.Button(root, text="GENERAR INFORME", command=self.generar, bg="#00D4FF",
                  fg="#0D0D1A", font=("Courier", 11, "bold"), relief="flat",
                  cursor="hand2", padx=18, pady=7).pack(pady=(4, 0))

    def fila(self, etiqueta, variable, comando):
        frame = tk.Frame(self.root, bg="#0D0D1A")
        frame.pack(fill="x", padx=30, pady=4)

        tk.Label(frame, text=etiqueta, font=("Courier", 9), fg="#AAAAAA",
                 bg="#0D0D1A", width=26, anchor="w").pack(side="left")
        tk.Label(frame, textvariable=variable, font=("Courier", 8), fg="#666666",
                 bg="#0D0D1A", width=16, anchor="w").pack(side="left")
        tk.Button(frame, text="Elegir", command=comando, bg="#1A3A5C", fg="#00D4FF",
                  font=("Courier", 9), relief="flat", cursor="hand2", padx=8).pack(side="left", padx=(6, 0))

    def elegir_bd(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar base de datos",
            filetypes=[("SQLite DB", "*.db"), ("Todos", "*.*")]
        )
        if ruta:
            self._bd = ruta
            self.ruta_bd.set(os.path.basename(ruta))

    def elegir_salida(self):
        carpeta = filedialog.askdirectory(title="Carpeta donde guardar el Excel")
        if carpeta:
            self._salida = carpeta
            self.ruta_salida.set(os.path.basename(carpeta))

    def generar(self):
        bd = getattr(self, "_bd", None)
        salida = getattr(self, "_salida", None)

        if not bd:
            messagebox.showwarning("Falta archivo", "Selecciona la base de datos primero.")
            return
        if not salida:
            messagebox.showwarning("Falta carpeta", "Elige dónde guardar el Excel.")
            return

        ruta_excel = os.path.join(salida, "informe_analisis.xlsx")

        try:
            self.lbl_estado.config(text="Generando...")
            self.root.update()
            generar_excel(bd, ruta_excel)
            self.lbl_estado.config(text="Guardado: informe_analisis.xlsx")
            messagebox.showinfo("Listo", f"Informe generado en:\n{ruta_excel}")
        except Exception as e:
            self.lbl_estado.config(text="Error al generar")
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()